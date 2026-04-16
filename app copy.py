# =====================================================================
# ADIDAS FINANCIAL ANALYSIS DASHBOARD — v2.0 Interactive Edition
# =====================================================================
# Three major interactivity upgrades over v1:
#   1. Metric Card Expanders   — click any metric card to drill into
#      formula breakdown, all-year trend chart, and thresholds
#   2. Drillable Balance Sheet — click any row in the table to open
#      a full detail panel: sparkline, % of total, YoY arrows, and
#      a list of which computed ratios are impacted by that item
#   3. Scenario Scatter Plot   — WC vs WCN journey map across 4 years
#      Click any year dot to open a full drill-down: what moved,
#      by how much, and what risk it implies
# =====================================================================

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import numpy as np
from pathlib import Path

# ─── Page configuration ─────────────────────────────────────────────
st.set_page_config(
    layout="wide",
    page_title="Adidas Financial Dashboard",
    page_icon="👟",
    initial_sidebar_state="expanded",
)

# ─── Constants ───────────────────────────────────────────────────────
FILE_PATH = Path(__file__).parent / "Easy-to-use simplified balance sheet Financial diagnosis_Adidas.xlsx"
YEARS = [2021, 2022, 2023, 2024]
YEARS_STR = [str(y) for y in YEARS]

BG      = "#0d1117"
CARD_BG = "#161b22"
BORDER  = "#30363d"
TXT     = "#ffffff"
TXT2    = "#8b949e"
BLUE    = "#004B87"
GREEN   = "#00A651"
ORANGE  = "#FF6B00"
RED     = "#E31937"
PALETTE = [BLUE, GREEN, ORANGE, RED, "#9B59B6", "#1ABC9C", "#F39C12", "#2ECC71"]

ASSET_ITEMS = [
    "Intangible assets", "Property, plant and equipment",
    "Other non-current assets", "Total non-current assets",
    "Inventories", "Accounts receivable", "Other current assets",
    "Cash and cash equivalents", "Total current assets", "Total assets",
]
EQLIAB_ITEMS = [
    "Issued share capital", "Capital & Other reserves",
    "Equity portion of convertible debt", "Retained earnings", "Total equity",
    "Borrowings non-current", "Other non-current liabilities",
    "Total non-current liabilities", "Accounts payable", "Bank overdraft",
    "Borrowings current", "Other current liabilities",
    "Total current liabilities", "Total equity and liabilities",
]
ALL_ITEMS = ASSET_ITEMS + EQLIAB_ITEMS

_SN = {
    "Property, plant and equipment": "PP&E",
    "Other non-current assets": "Other NCA",
    "Total non-current assets": "Total NCA",
    "Cash and cash equivalents": "Cash & Equiv.",
    "Total current assets": "Total CA",
    "Accounts receivable": "Accounts Recv.",
    "Other current assets": "Other CA",
    "Issued share capital": "Share Capital",
    "Capital & Other reserves": "Capital & Reserves",
    "Equity portion of convertible debt": "Conv. Debt Equity",
    "Borrowings non-current": "Borrowings (NC)",
    "Other non-current liabilities": "Other NCL",
    "Total non-current liabilities": "Total NCL",
    "Accounts payable": "Accounts Pay.",
    "Borrowings current": "Borrowings (Curr.)",
    "Other current liabilities": "Other CL",
    "Total current liabilities": "Total CL",
    "Total equity and liabilities": "Total E&L",
}
def sn(n): return _SN.get(n, n)

# ─── Scenario definitions ────────────────────────────────────────────
SCENARIOS = {
    1: ("Ideal",          "✅", GREEN,    "WC > 0 · WCN > 0 · NC > 0",
        "Healthy cycle. Long-term resources cover fixed assets with extra cash for operations."),
    2: ("Risky but Common","⚠️", ORANGE,  "WC > 0 · WCN > 0 · NC < 0",
        "Stable long-term, but short-term liquidity tension. Operating cycle too heavy, needs short-term bank loans."),
    3: ("Excellent",       "🌟","#00D4AA","WC > 0 · WCN < 0 · NC > 0",
        "Perfect situation (common in retail). Negative WCN means operations generate cash. Solid long-term, excess liquidity."),
    4: ("Paradoxical",     "🔄","#FFD700","WC < 0 · WCN < 0 · NC > 0",
        "Risky structure (fixed assets financed by short-term debt), but business generates enough daily cash to survive. If activity slows, massive liquidity problem."),
    5: ("Risky",           "🔴", RED,     "WC < 0 · WCN < 0 · NC < 0",
        "Bad structure + cash consumption. Extreme financial tension, over-reliant on short-term debt."),
    6: ("Dangerous",       "💀","#FF0000","WC < 0 · WCN > 0 · NC < 0",
        "Worst situation. Weak structure, cash-consuming operating cycle, severe liquidity distress. High bankruptcy risk."),
}

# ─── Balance-sheet item → impacted metrics (for drill panel) ────────
ITEM_IMPACTS = {
    "Inventories": {
        "metrics": ["WCN ↑", "Quick Ratio ↓", "Current Ratio (indirect)"],
        "note": "Higher inventories raise WCN (more cash tied up in operations) and are excluded from the Quick Ratio."
    },
    "Cash and cash equivalents": {
        "metrics": ["Net Cash (NC) ↑", "Cash Ratio ↑", "Net Debt ↓"],
        "note": "More cash directly improves Net Cash, Cash Ratio, and reduces Net Debt."
    },
    "Accounts receivable": {
        "metrics": ["WCN ↑", "Current Ratio (indirect)"],
        "note": "Higher receivables increase WCN — more cash is tied up waiting to be collected."
    },
    "Other current assets": {
        "metrics": ["WCN ↑", "Current Ratio (indirect)"],
        "note": "Part of the operating working capital need formula."
    },
    "Total equity": {
        "metrics": ["WC ↑", "Equity Multiplier ↓", "Debt-to-Equity ↓", "Invested Capital"],
        "note": "Larger equity base improves WC and reduces all leverage ratios."
    },
    "Retained earnings": {
        "metrics": ["Total equity → WC", "Equity Multiplier ↓", "Debt-to-Equity ↓"],
        "note": "Retained earnings build the equity base over time, reducing leverage."
    },
    "Borrowings non-current": {
        "metrics": ["WC ↑ (part of long-term resources)", "Net Debt ↑", "Debt-to-Equity ↑", "Total Debt Ratio ↑"],
        "note": "Long-term borrowings add to WC resources but raise financial leverage."
    },
    "Borrowings current": {
        "metrics": ["Net Cash (NC) ↓", "Net Debt ↑", "Debt-to-Equity ↑", "Total Debt Ratio ↑"],
        "note": "Short-term financial debt directly reduces Net Cash and raises all debt ratios."
    },
    "Accounts payable": {
        "metrics": ["WCN ↓ (supplier credit reduces need)"],
        "note": "Higher payables lower WCN — suppliers are effectively financing part of operations."
    },
    "Other current liabilities": {
        "metrics": ["WCN ↓", "Current Ratio ↓"],
        "note": "Non-financial current liabilities subtract from WCN and are in the Current Ratio denominator."
    },
    "Total non-current assets": {
        "metrics": ["WC ↓ (higher fixed assets consume long-term resources)", "Capital Employed ↑"],
        "note": "More fixed assets reduce WC and increase Capital Employed."
    },
    "Total non-current liabilities": {
        "metrics": ["WC ↑ (part of long-term resources)"],
        "note": "Long-term liabilities are permanent resources that help fund fixed assets."
    },
    "Property, plant and equipment": {
        "metrics": ["Total NCA → WC ↓", "Capital Employed ↑"],
        "note": "PP&E is the largest fixed asset component, a key driver of WC and Capital Employed."
    },
    "Intangible assets": {
        "metrics": ["Total NCA → WC ↓"],
        "note": "Intangibles (brands, patents) are part of non-current assets."
    },
    "Issued share capital": {
        "metrics": ["Total equity → WC", "Equity Multiplier ↓"],
        "note": "Represents paid-in capital at par value."
    },
}

# ─── CSS injection ───────────────────────────────────────────────────
def inject_css():
    st.markdown(f"""<style>
    .stApp {{ background:{BG}; color:{TXT}; }}
    [data-testid="stSidebar"] {{ background:#0a0e14; border-right:1px solid {BORDER}; }}
    [data-testid="stSidebar"] * {{ color:{TXT}!important; }}
    /* cards */
    .card {{ background:{CARD_BG}; border:1px solid {BORDER}; border-radius:12px;
             padding:18px 20px; margin-bottom:10px; }}
    .card h4 {{ color:{TXT2}; font-size:.82rem; margin:0 0 4px; font-weight:500; }}
    .card .val {{ font-size:1.85rem; font-weight:700; line-height:1.2; }}
    .card .sub {{ color:{TXT2}; font-size:.78rem; margin-top:4px; }}
    /* scenario cards */
    .scenario {{ background:{CARD_BG}; border:1px solid {BORDER}; border-radius:12px;
                 padding:14px 18px; margin-bottom:10px; border-left:4px solid; }}
    /* interpretation panels */
    .ipanel {{ background:{CARD_BG}; border:1px solid {BORDER}; border-radius:10px;
               padding:16px; margin:8px 0; }}
    .ipanel .formula {{ font-family:'Courier New',monospace; background:{BG}; padding:8px 12px;
                        border-radius:6px; margin:6px 0; color:{BLUE}; font-size:.85rem; }}
    .ipanel .verdict {{ margin-top:8px; padding:10px; border-radius:6px; font-weight:600; }}
    /* section headers */
    .shdr {{ font-size:1.15rem; font-weight:700; margin:18px 0 8px;
             padding-bottom:6px; border-bottom:2px solid {BLUE}; color:{TXT}; }}
    /* drill panel */
    .drill-panel {{ background:{CARD_BG}; border:1px solid {BLUE}; border-radius:12px;
                    padding:16px; margin:12px 0; }}
    .impact-tag {{ display:inline-block; background:{BG}; border:1px solid {BORDER};
                   border-radius:6px; padding:3px 8px; margin:2px; font-size:.78rem; color:{BLUE}; }}
    footer {{ visibility:hidden; }}
    </style>""", unsafe_allow_html=True)


# ─── Data loading ────────────────────────────────────────────────────
@st.cache_data
def load_data():
    import openpyxl
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    ws = wb["Balance Sheet"]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    header = rows[0]
    ycols = {v: i for i, v in enumerate(header) if v in YEARS}
    data, bcount = {}, 0
    for r in rows[1:]:
        lbl = r[1]
        if lbl is None:
            continue
        if "Borrowings" in str(lbl) and "similar" in str(lbl):
            bcount += 1
            lbl = "Borrowings non-current" if bcount == 1 else "Borrowings current"
        data[lbl] = {y: (r[c] if r[c] is not None else 0) for y, c in ycols.items()}
    df = pd.DataFrame(data).T[sorted(ycols.keys())].astype(float)
    return df  # €K


# ─── Computations ────────────────────────────────────────────────────
def _g(df, label, yr):
    try:   return float(df.loc[label, yr])
    except: return 0.0

def compute(df):
    out = {}
    for yr in YEARS:
        eq=_g(df,"Total equity",yr); ncl=_g(df,"Total non-current liabilities",yr)
        nca=_g(df,"Total non-current assets",yr); ca=_g(df,"Total current assets",yr)
        cl=_g(df,"Total current liabilities",yr); ta=_g(df,"Total assets",yr)
        inv=_g(df,"Inventories",yr); ar=_g(df,"Accounts receivable",yr)
        oca=_g(df,"Other current assets",yr); cash=_g(df,"Cash and cash equivalents",yr)
        ap=_g(df,"Accounts payable",yr); ocl=_g(df,"Other current liabilities",yr)
        bnc=_g(df,"Borrowings non-current",yr); bc=_g(df,"Borrowings current",yr)
        bo=_g(df,"Bank overdraft",yr)
        wc=eq+ncl-nca; wcn=(inv+ar+oca)-(ap+ocl); nc=wc-wcn
        nd=ncl+bc-cash; ce=nca+wcn; ic=eq+nd; debt=ncl+bc
        cr=ca/cl if cl else 0; qr=(ca-inv)/cl if cl else 0
        cashr=cash/cl if cl else 0; em=ta/eq if eq else 0
        tdr=debt/ta if ta else 0; de=debt/eq if eq else 0
        s=(1 if wc>0 and wcn>0 and nc>0 else 2 if wc>0 and wcn>0 and nc<0 else
           3 if wc>0 and wcn<0 and nc>0 else 4 if wc<0 and wcn<0 and nc>0 else
           5 if wc<0 and wcn<0 and nc<0 else 6 if wc<0 and wcn>0 and nc<0 else 0)
        out[yr]={"WC":wc,"WCN":wcn,"NC":nc,"NC_verify":cash-bc-bo,
                 "Net Debt":nd,"Capital Employed":ce,"Invested Capital":ic,
                 "Current Ratio":cr,"Quick Ratio":qr,"Cash Ratio":cashr,
                 "Equity Multiplier":em,"Total Debt Ratio":tdr,"Debt-to-Equity":de,
                 "Scenario":s}
    return out


# ─── Rating functions ────────────────────────────────────────────────
def rate_cr(v):
    if v<1:    return "🔴",RED,"Liquidity concern — current liabilities exceed current assets"
    if v<1.5:  return "🟡",ORANGE,"Below ideal range — adequate but tight"
    if v<=2:   return "🟢",GREEN,"Healthy — adequate liquidity for short-term obligations"
    if v<=3:   return "🟢",GREEN,"Good liquidity position"
    return "🟡",ORANGE,"Inefficient — too much working capital tied up"

def rate_qr(v):
    if v<1:    return "🔴",RED,"Liquidity risk — heavily dependent on selling inventory"
    if v<=1.5: return "🟢",GREEN,"Strong, safe liquidity without relying on inventory"
    if v<=2:   return "🟢",GREEN,"Very good quick liquidity"
    return "🟡",ORANGE,"Very conservative — could mean under-investing in growth"

def rate_cashr(v):
    if v<0.5:  return "🟡",ORANGE,"Reliance on inventory/receivables to pay short-term debts"
    if v<=1:   return "🟢",GREEN,"Ideal cash position relative to current liabilities"
    return "🟡",ORANGE,"Extremely conservative — excess cash may be idle"

def rate_tdr(v):
    if v<=0.2: return "🟢",GREEN,"Very low debt"
    if v<=0.4: return "🟢",GREEN,"Low debt — healthy capital structure"
    if v<=0.6: return "🟡",ORANGE,"Acceptable for most industries"
    if v<=0.8: return "🔴",RED,"High debt load — caution advised"
    return "🔴🔴",RED,"Very high leverage — significant solvency risk"

def rate_de(v):
    if v<=0.5: return "🟢",GREEN,"Very healthy — low leverage"
    if v<=1:   return "🟢",GREEN,"Healthy, balanced capital structure"
    if v<=1.5: return "🟡",ORANGE,"Acceptable leverage"
    return "🔴",RED,"High leverage — potential solvency concerns"

def rate_sign(v, label=""):
    if v>0: return "🟢",GREEN,f"Positive {label}".strip()
    if v<0: return "🔴",RED,  f"Negative {label}".strip()
    return "⚪",TXT2,f"Zero {label}".strip()


# ─── Interpretation helpers ──────────────────────────────────────────

def _trend(vals_dict):
    """Overall trend direction across 4 years."""
    vals = [vals_dict[yr] for yr in YEARS]
    diffs = [vals[i+1] - vals[i] for i in range(3)]
    pos = sum(1 for d in diffs if d > 0)
    neg = sum(1 for d in diffs if d < 0)
    if pos == 3: return "consistently improving"
    if neg == 3: return "consistently declining"
    if pos == 2: return "mostly improving"
    if neg == 2: return "mostly declining"
    return "volatile"

def _yr_label_cr(v):
    if v < 1:   return "🔴","Risky",                f"({fr(v)}) — may struggle to cover short-term obligations"
    if v < 1.5: return "🟡","Adequate but tight",   f"({fr(v)}) — below the healthy 1.5–2.0 range"
    if v <= 2:  return "🟢","Healthy",              f"({fr(v)}) — {fr(v)}× assets vs short-term liabilities"
    if v <= 3:  return "🟢","Good",                 f"({fr(v)}) — solid short-term liquidity cushion"
    return      "🟡","Potentially inefficient",     f"({fr(v)}) — excess working capital may be idle"

def _yr_label_qr(v):
    if v < 1:   return "🔴","Risky",              f"({fr(v)}) — may rely on inventory or borrowing to meet short-term debts"
    if v <= 1.5:return "🟢","Strong liquidity",   f"({fr(v)}) — covers liabilities without relying on inventory"
    if v <= 2:  return "🟢","Very good",          f"({fr(v)}) — highly liquid even without inventory"
    return      "🟡","Very conservative",         f"({fr(v)}) — may signal under-investment in growth"

def _yr_label_cashr(v):
    if v < 0.5: return "🟡","Below ideal",           f"({fr(v)}) — relies on receivables/inventory; not necessarily bad"
    if v <= 1:  return "🟢","Ideal",                 f"({fr(v)}) — covers {int(v*100)}% of short-term debts immediately"
    return      "🟡","Extremely conservative",       f"({fr(v)}) — more cash than all short-term liabilities"

def _yr_label_tdr(v):
    if v <= 0.2:return "🟢","Very low debt",         f"({fr(v)}) — highly equity-financed"
    if v <= 0.4:return "🟢","Low–moderate debt",     f"({fr(v)}) — solid solvency; most assets financed by equity"
    if v <= 0.6:return "🟡","Acceptable",            f"({fr(v)}) — some exposure to downturns"
    if v <= 0.8:return "🔴","High debt load",        f"({fr(v)}) — vulnerability warning signal"
    return      "🔴","Very high leverage",           f"({fr(v)}) — severe solvency risk"

def _yr_label_de(v):
    if v <= 0.5:return "🟢","Very healthy",          f"({fr(v)}) — low leverage"
    if v <= 1:  return "🟢","Healthy",               f"({fr(v)}) — balanced capital structure"
    if v <= 1.5:return "🟡","Acceptable",            f"({fr(v)}) — normal for stable industries"
    if v <= 2:  return "🔴","High leverage",         f"({fr(v)}) — more debt than equity"
    return      "🔴","Very high leverage",           f"({fr(v)}) — potential solvency problems"

def _yr_label_em(v):
    if v <= 1.5:return "🟢","Low leverage",          f"({fr(v)}) — mostly equity-financed"
    if v <= 2.5:return "🟡","Moderate leverage",     f"({fr(v)}) — balanced debt/equity mix"
    if v <= 4:  return "🔴","High leverage",         f"({fr(v)}) — aggressive financial management"
    return      "🔴","Very high leverage",           f"({fr(v)}) — heavy reliance on debt"

def _yr_label_wc(v):
    if v > 0:   return "🟢","Positive",  f"({fm(v)}) — long-term resources exceed fixed assets → financial safety margin"
    if v < 0:   return "🔴","Negative",  f"({fm(v)}) — fixed assets partly financed by short-term debt → structurally risky"
    return      "⚪","Zero",             "— long-term resources exactly cover fixed assets"

def _yr_label_wcn(v):
    if v > 0:   return "🔴","Positive",  f"({fm(v)}) — operating cycle needs financing; cash tied up in operations"
    if v < 0:   return "🟢","Negative",  f"({fm(v)}) — operating cycle self-financed; suppliers fund the business"
    return      "⚪","Zero",             "— operating cycle is self-financing"

def _yr_label_nc(v):
    if v > 0:   return "🟢","Positive",  f"({fm(v)}) — cash surplus after financing the operating cycle"
    if v < 0:   return "🔴","Negative",  f"({fm(v)}) — operating cycle consumes more than long-term cushion → needs short-term financing"
    return      "⚪","Zero",             "— WC exactly covers WCN"

def _yr_label_nd(v):
    if v > 0:   return "🔴","Net Debt",  f"of {fm(v)} — more obligations than cash → leverage risk"
    if v < 0:   return "🟢","Net Cash",  f"of {fm(abs(v))} — cash exceeds all obligations → strong financial stability"
    return      "⚪","Neutral",          "— cash exactly offsets debt"

def _yr_label_ce(v):
    return "🔵","Capital deployed", f"of {fm(v)} — total capital employed in operations (NCA + WCN)"

def _yr_label_ic(v):
    return "🔵","Capital invested", f"of {fm(v)} — total invested capital (Equity + Net Debt), equals Capital Employed"

def _evo_html(vals_dict, yr_label_fn, metric_name, higher_is_better=True):
    v0, v3 = vals_dict[YEARS[0]], vals_dict[YEARS[-1]]
    trend = _trend(vals_dict)
    chg = yoy_pct(v3, v0)
    chg_str = f" ({abs(chg):.1f}% {'higher' if chg >= 0 else 'lower'} than 2021)" if chg is not None else ""
    good = ("improving" in trend and higher_is_better) or ("declining" in trend and not higher_is_better)
    bad  = ("declining" in trend and higher_is_better) or ("improving" in trend and not higher_is_better)
    color = GREEN if good else RED if bad else ORANGE
    ico0, lbl0, _ = yr_label_fn(v0)
    ico3, lbl3, _ = yr_label_fn(v3)
    return (
        f'<div style="margin-top:8px;padding:8px 10px;background:{BG};border-radius:6px;font-size:.82rem;">'
        f'<span style="color:{TXT2};">📈 2021→2024: </span>'
        f'<span style="color:{color};font-weight:600;">{trend}</span>{chg_str}. '
        f'<span style="color:{TXT2};">Moved from </span>{ico0} <b>{lbl0}</b>'
        f'<span style="color:{TXT2};"> in 2021 → </span>{ico3} <b>{lbl3}</b>'
        f'<span style="color:{TXT2};"> in 2024.</span>'
        f'</div>'
    )

def _interp_block(vals_dict, yr_label_fn, metric_name, higher_is_better=True):
    """Full interpretation block: per-year table + evolution summary."""
    rows = ""
    for yr in YEARS:
        ico, lbl, detail = yr_label_fn(vals_dict[yr])
        rows += (
            f'<tr>'
            f'<td style="padding:3px 10px 3px 0;color:{TXT2};font-size:.78rem;'
            f'white-space:nowrap;vertical-align:top;">{yr}</td>'
            f'<td style="padding:3px 0;font-size:.82rem;line-height:1.4;">'
            f'{ico} <b>{lbl}</b> {detail}</td>'
            f'</tr>'
        )
    evo = _evo_html(vals_dict, yr_label_fn, metric_name, higher_is_better)
    return (
        f'<div style="margin-top:6px;">'
        f'<div style="font-size:.78rem;font-weight:600;color:{TXT2};'
        f'text-transform:uppercase;letter-spacing:.05em;margin-bottom:4px;">📖 Interpretation</div>'
        f'<table style="width:100%;border-collapse:collapse;">{rows}</table>'
        + evo +
        f'</div>'
    )


# ─── Formatting helpers ──────────────────────────────────────────────
def fm(v_k):  return f"€{v_k/1000:,.1f}M"
def fr(v):    return f"{v:.2f}"
def yoy_pct(curr,prev): return ((curr-prev)/abs(prev)*100) if prev!=0 else None

def arrow_html(val):
    if val is None: return f'<span style="color:{TXT2};">N/A</span>'
    c = GREEN if val>=0 else RED
    s = "↑" if val>=0 else "↓"
    return f'<span style="color:{c};font-weight:600;">{s}&nbsp;{abs(val):.1f}%</span>'


# ─── Chart helper ────────────────────────────────────────────────────
def dark_fig(fig, height=420):
    fig.update_layout(
        template="plotly_dark", paper_bgcolor=BG, plot_bgcolor=BG,
        font=dict(family="sans-serif",color=TXT,size=12),
        legend=dict(bgcolor="rgba(0,0,0,0)",bordercolor=BORDER),
        margin=dict(l=50,r=30,t=50,b=40), height=height,
    )
    fig.update_xaxes(gridcolor=BORDER,linecolor=BORDER)
    fig.update_yaxes(gridcolor=BORDER,linecolor=BORDER)
    return fig

def mini_trend_chart(y_vals, y_labels, title, color, is_eur, height=200, key=""):
    """Small bar chart used inside expanders."""
    bar_colors = [GREEN if v>=0 else RED for v in y_vals] if is_eur else [color]*len(y_vals)
    texts = [fm(v*1000) if is_eur else fr(v) for v in y_vals]
    fig = go.Figure(go.Bar(
        x=y_labels, y=y_vals, marker_color=bar_colors,
        text=texts, textposition="outside", textfont=dict(color=TXT,size=10)
    ))
    if is_eur:
        fig.add_hline(y=0, line_dash="dash", line_color=TXT2, opacity=0.3)
    fig.update_layout(
        template="plotly_dark", paper_bgcolor=BG, plot_bgcolor=BG,
        margin=dict(l=10,r=10,t=30,b=20), height=height,
        title=dict(text=title,font=dict(size=11,color=TXT2)),
        showlegend=False,
        yaxis_title="€M" if is_eur else "",
    )
    fig.update_xaxes(gridcolor=BORDER); fig.update_yaxes(gridcolor=BORDER)
    return fig


# ─── UI primitives (kept for places that don't need expanders) ───────
def card_html(title, value, color=TXT, subtitle=""):
    st.markdown(
        f'<div class="card"><h4>{title}</h4>'
        f'<div class="val" style="color:{color};">{value}</div>'
        f'<div class="sub">{subtitle}</div></div>',
        unsafe_allow_html=True)

def scenario_card_html(year, case_num):
    label,icon,color,signs,interp = SCENARIOS.get(case_num,("Unknown","❓",TXT2,"",""))
    st.markdown(
        f'<div class="scenario" style="border-left-color:{color};">'
        f'<div style="font-size:.78rem;color:{TXT2};">FY {year}</div>'
        f'<div style="font-size:1.3rem;font-weight:700;color:{color};">'
        f'{icon} Case {case_num}: {label}</div>'
        f'<div style="font-size:.72rem;color:{TXT2};margin:4px 0;">{signs}</div>'
        f'<div style="font-size:.82rem;color:{TXT};margin-top:6px;">{interp}</div></div>',
        unsafe_allow_html=True)

def interp_panel_html(title, formula_text, value_str, thresholds_html, verdict, vcolor):
    st.markdown(
        f'<div class="ipanel">'
        f'<div style="font-weight:700;font-size:1rem;margin-bottom:6px;">{title}</div>'
        f'<div class="formula">{formula_text}</div>'
        f'<div style="margin:6px 0;"><strong>Value:</strong> '
        f'<span style="color:{vcolor};font-size:1.2rem;font-weight:700;">{value_str}</span></div>'
        f'<div style="margin:6px 0;font-size:.85rem;">{thresholds_html}</div>'
        f'<div class="verdict" style="background:{vcolor}18;border-left:3px solid {vcolor};color:{vcolor};">'
        f'{verdict}</div></div>', unsafe_allow_html=True)

def section(title):
    st.markdown(f'<div class="shdr">{title}</div>', unsafe_allow_html=True)


# =====================================================================
# ★ FEATURE 1 — Interactive Metric Card Expander
#   Click the card → formula, all-year values, mini trend chart,
#   thresholds, and impacted metrics
# =====================================================================
def metric_card_expander(
    title, all_year_vals, is_eur, formula,
    threshold_html=None, impacts=None, key_prefix="card",
    rate_fn=None, interpret_fn=None,
):
    """
    Interactive card. Shows 2024 value as the expander label.
    Expands to reveal: all-year table, mini trend chart, formula,
    thresholds, and which other metrics this one affects.

    all_year_vals : dict {year: numeric value in €K if is_eur, else raw ratio}
    is_eur        : True → format as €M (divide by 1000), False → ratio
    rate_fn       : optional function(value) → (emoji, color, text) for per-year verdict
    """
    v24 = all_year_vals[2024]
    if is_eur:
        display_val = fm(v24)
        _, color24, sub24 = rate_sign(v24)
    else:
        display_val = fr(v24)
        if rate_fn:
            _, color24, sub24 = rate_fn(v24)
        else:
            color24, sub24 = BLUE, ""

    expander_label = f"**{display_val}** — {title}"

    with st.expander(expander_label, expanded=False):
        # ── 2024 highlight ──
        st.markdown(
            f'<div style="color:{color24};font-size:1.6rem;font-weight:700;">{display_val}</div>'
            f'<div style="color:{TXT2};font-size:.8rem;margin-bottom:8px;">{sub24}</div>',
            unsafe_allow_html=True)

        # ── All-years mini table ──
        cols = st.columns(4)
        for i, yr in enumerate(YEARS):
            v = all_year_vals[yr]
            dv = fm(v) if is_eur else fr(v)
            if is_eur:
                _, c, sub = rate_sign(v)
            else:
                if rate_fn:
                    _, c, sub = rate_fn(v)
                else:
                    c, sub = BLUE, ""
            with cols[i]:
                st.markdown(
                    f'<div style="background:{BG};border:1px solid {BORDER};border-radius:8px;'
                    f'padding:8px 10px;text-align:center;">'
                    f'<div style="font-size:.7rem;color:{TXT2};">{yr}</div>'
                    f'<div style="font-size:1rem;font-weight:700;color:{c};">{dv}</div>'
                    f'</div>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Mini trend chart ──
        y_vals_display = [all_year_vals[yr]/1000 if is_eur else all_year_vals[yr] for yr in YEARS]
        unit = "€M" if is_eur else ""
        fig = mini_trend_chart(y_vals_display, YEARS_STR, f"{title} — 4-year trend",
                               color24, is_eur, height=190,
                               key=f"{key_prefix}_mini")
        st.plotly_chart(fig, use_container_width=True, key=f"{key_prefix}_mini_chart")

        # ── YoY changes ──
        yoy_parts = []
        for i in range(1, len(YEARS)):
            chg = yoy_pct(all_year_vals[YEARS[i]], all_year_vals[YEARS[i-1]])
            yoy_parts.append(f"{YEARS[i]}: {arrow_html(chg)}")
        st.markdown(" &nbsp;|&nbsp; ".join(yoy_parts), unsafe_allow_html=True)

        st.markdown("<hr style='border-color:{};margin:8px 0;'>".format(BORDER),
                    unsafe_allow_html=True)

        # ── Formula ──
        if formula:
            st.markdown(
                f'<div style="font-family:monospace;background:{BG};padding:8px 12px;'
                f'border-radius:6px;color:{BLUE};font-size:.82rem;margin:6px 0;">'
                f'{formula}</div>', unsafe_allow_html=True)

        # ── Thresholds ──
        if threshold_html:
            st.markdown(
                f'<div style="font-size:.84rem;margin:6px 0;">{threshold_html}</div>',
                unsafe_allow_html=True)

        # ── Interpretation ──
        if interpret_fn:
            st.markdown(interpret_fn(all_year_vals), unsafe_allow_html=True)

        # ── Impacted metrics ──
        if impacts:
            tags = "".join(f'<span class="impact-tag">{m}</span>' for m in impacts)
            st.markdown(
                f'<div style="margin-top:8px;"><span style="font-size:.78rem;color:{TXT2};">'
                f'Impacts:</span><br>{tags}</div>', unsafe_allow_html=True)


# =====================================================================
# ★ FEATURE 2 — Drillable Balance Sheet helpers
# =====================================================================
def bs_drill_panel(df, M, selected_item):
    """
    Full drill-down panel for one balance sheet line item.
    Shows: sparkline, % of total, YoY changes, impacted metrics.
    """
    if selected_item is None:
        return

    is_asset = selected_item in ASSET_ITEMS
    denom_key = "Total assets" if is_asset else "Total equity and liabilities"

    vals_k = [_g(df, selected_item, yr) for yr in YEARS]
    vals_m = [v/1000 for v in vals_k]
    denoms = [_g(df, denom_key, yr) for yr in YEARS]
    pcts   = [v/d*100 if d else 0 for v, d in zip(vals_k, denoms)]

    st.markdown(
        f'<div class="drill-panel">'
        f'<div style="font-size:1.1rem;font-weight:700;color:{BLUE};margin-bottom:10px;">'
        f'🔍 Deep Dive — {selected_item}</div></div>',
        unsafe_allow_html=True)

    c1, c2 = st.columns([2, 3])

    with c1:
        # Per-year cards
        for yr, vm, pct in zip(YEARS, vals_m, pcts):
            chg = yoy_pct(_g(df, selected_item, yr), _g(df, selected_item, YEARS[YEARS.index(yr)-1])) \
                  if yr != YEARS[0] else None
            delta_str = f" ({arrow_html(chg)})" if chg is not None else ""
            st.markdown(
                f'<div style="background:{BG};border:1px solid {BORDER};border-radius:8px;'
                f'padding:8px 12px;margin-bottom:6px;display:flex;justify-content:space-between;">'
                f'<span style="color:{TXT2};font-size:.82rem;">{yr}</span>'
                f'<span style="font-weight:700;">€{vm:,.1f}M</span>'
                f'<span style="color:{BLUE};font-size:.82rem;">{pct:.1f}% of {sn(denom_key)}</span>'
                f'{delta_str}</div>',
                unsafe_allow_html=True)

    with c2:
        # Trend chart + % of total chart
        tab1, tab2 = st.tabs(["📈 Trend (€M)", "📊 % of Total"])
        with tab1:
            fig = go.Figure()
            bar_colors = [GREEN if v>=0 else RED for v in vals_m]
            fig.add_trace(go.Bar(
                x=YEARS_STR, y=vals_m, marker_color=bar_colors,
                text=[f"€{v:,.1f}M" for v in vals_m], textposition="outside",
                textfont=dict(color=TXT, size=11)
            ))
            fig.add_trace(go.Scatter(
                x=YEARS_STR, y=vals_m, mode="lines+markers",
                line=dict(color=BLUE, width=2, dash="dot"),
                marker=dict(size=7), name="trend", showlegend=False
            ))
            fig.update_layout(
                template="plotly_dark", paper_bgcolor=BG, plot_bgcolor=BG,
                margin=dict(l=10,r=10,t=30,b=20), height=250,
                showlegend=False, yaxis_title="€M",
                title=dict(text=f"{sn(selected_item)} — Values (€M)", font=dict(size=11))
            )
            fig.update_xaxes(gridcolor=BORDER); fig.update_yaxes(gridcolor=BORDER)
            st.plotly_chart(fig, use_container_width=True, key=f"drill_trend_{selected_item}")

        with tab2:
            fig2 = go.Figure(go.Bar(
                x=YEARS_STR, y=pcts, marker_color=PALETTE[:4],
                text=[f"{p:.1f}%" for p in pcts], textposition="outside",
                textfont=dict(color=TXT, size=11)
            ))
            fig2.update_layout(
                template="plotly_dark", paper_bgcolor=BG, plot_bgcolor=BG,
                margin=dict(l=10,r=10,t=30,b=20), height=250,
                showlegend=False, yaxis_title=f"% of {sn(denom_key)}",
                title=dict(text=f"As % of {sn(denom_key)}", font=dict(size=11))
            )
            fig2.update_xaxes(gridcolor=BORDER); fig2.update_yaxes(gridcolor=BORDER)
            st.plotly_chart(fig2, use_container_width=True, key=f"drill_pct_{selected_item}")

    # Impacted metrics
    impact_info = ITEM_IMPACTS.get(selected_item)
    if impact_info:
        tags = "".join(f'<span class="impact-tag">{m}</span>' for m in impact_info["metrics"])
        st.markdown(
            f'<div style="margin-top:8px;">'
            f'<div style="font-size:.82rem;color:{TXT2};margin-bottom:4px;">Impacted computed metrics:</div>'
            f'{tags}<br>'
            f'<div style="font-size:.8rem;color:{TXT2};margin-top:6px;">💡 {impact_info["note"]}</div>'
            f'</div>', unsafe_allow_html=True)


# =====================================================================
# ★ FEATURE 3 — Scenario Scatter Plot (WC vs WCN journey map)
# =====================================================================
def scenario_scatter_section(M):
    """
    Plots Adidas' WC vs WCN position over 4 years.
    The diagonal WC = WCN marks the NC = 0 boundary.
    Background shading marks Case 1 (NC>0) and Case 2 (NC<0) zones.
    Clicking a year point shows a drill-down below the chart.
    """
    section("📍 Scenario Position Map — Click a Year to Drill Down")
    st.caption(
        "X = Working Capital (€M) · Y = Working Capital Need (€M) · "
        "The dashed diagonal (WC = WCN) separates NC > 0 (below) from NC < 0 (above)."
    )

    wc  = {yr: M[yr]["WC"] /1000 for yr in YEARS}
    wcn = {yr: M[yr]["WCN"]/1000 for yr in YEARS}
    nc  = {yr: M[yr]["NC"] /1000 for yr in YEARS}

    all_wc  = list(wc.values())
    all_wcn = list(wcn.values())
    pad = 500
    x_min = min(all_wc)  - pad;  x_max = max(all_wc)  + pad
    y_min = min(all_wcn) - pad;  y_max = max(all_wcn) + pad
    diag_range = max(x_max, y_max) + 200

    fig = go.Figure()

    # ── Background shading: NC > 0 zone (below diagonal, Case 1) ──
    fig.add_trace(go.Scatter(
        x=[0, diag_range, diag_range, 0],
        y=[0, diag_range, y_min,      y_min],
        fill="toself",
        fillcolor=f"rgba(0,166,81,0.07)",
        line=dict(width=0), hoverinfo="skip", showlegend=False, name=""
    ))

    # ── Background shading: NC < 0 zone (above diagonal, Case 2) ──
    fig.add_trace(go.Scatter(
        x=[0, 0,          diag_range, diag_range],
        y=[0, y_max+200,  y_max+200,  diag_range],
        fill="toself",
        fillcolor=f"rgba(255,107,0,0.07)",
        line=dict(width=0), hoverinfo="skip", showlegend=False, name=""
    ))

    # ── Diagonal: NC = 0 line (WC = WCN) ──
    fig.add_trace(go.Scatter(
        x=[0, diag_range], y=[0, diag_range],
        mode="lines", line=dict(color=TXT2, width=1.5, dash="dash"),
        name="NC = 0  (WC = WCN)", hoverinfo="skip"
    ))

    # ── Connecting path ──
    fig.add_trace(go.Scatter(
        x=[wc[yr] for yr in YEARS], y=[wcn[yr] for yr in YEARS],
        mode="lines", line=dict(color=TXT2, width=1, dash="dot"),
        name="Adidas path 2021→2024", hoverinfo="skip"
    ))

    # ── Year points (one trace each so on_select gives curve_number) ──
    for yr in YEARS:
        case_num = M[yr]["Scenario"]
        label, icon, color, signs, interp = SCENARIOS[case_num]
        nc_val = nc[yr]
        hover = (
            f"<b>FY {yr}</b><br>"
            f"WC: {fm(M[yr]['WC'])}<br>"
            f"WCN: {fm(M[yr]['WCN'])}<br>"
            f"NC: {fm(M[yr]['NC'])}<br>"
            f"━━━━━━━<br>"
            f"{icon} Case {case_num}: <b>{label}</b><br>"
            f"<i>{signs}</i>"
        )
        fig.add_trace(go.Scatter(
            x=[wc[yr]], y=[wcn[yr]],
            mode="markers+text",
            marker=dict(size=22, color=color, symbol="circle",
                        line=dict(color=TXT, width=2)),
            text=[f"<b>{yr}</b>"], textposition="top center",
            textfont=dict(color=TXT, size=12, family="sans-serif"),
            name=f"{yr}: {icon} {label}",
            hovertext=hover, hoverinfo="text",
            customdata=[yr],
        ))

    # ── Zone labels ──
    fig.add_annotation(
        x=x_max - 200, y=y_min + 200,
        text="① Ideal (NC > 0)", font=dict(color=GREEN, size=11),
        bgcolor="rgba(0,166,81,0.12)", bordercolor=GREEN,
        showarrow=False, borderwidth=1, borderpad=4
    )
    fig.add_annotation(
        x=x_min + 300, y=y_max - 200,
        text="② Risky but Common (NC < 0)", font=dict(color=ORANGE, size=11),
        bgcolor="rgba(255,107,0,0.12)", bordercolor=ORANGE,
        showarrow=False, borderwidth=1, borderpad=4
    )

    fig.update_layout(
        xaxis_title="Working Capital (€M)",
        yaxis_title="Working Capital Need (€M)",
        xaxis=dict(zeroline=True, zerolinecolor=TXT2, zerolinewidth=1.5,
                   range=[x_min, x_max]),
        yaxis=dict(zeroline=True, zerolinecolor=TXT2, zerolinewidth=1.5,
                   range=[y_min, y_max]),
        hovermode="closest",
    )
    dark_fig(fig, 520)

    # ── Render with on_select for click-to-drill ──
    event = st.plotly_chart(
        fig, use_container_width=True,
        on_select="rerun", key="scenario_scatter_map"
    )

    # ── Click-based drill-down ──
    # Year traces are indices 4..7 (after bg1, bg2, diagonal, path)
    YEAR_TRACE_OFFSET = 4
    selected_yr = None

    if event and hasattr(event, "selection") and event.selection.points:
        pt = event.selection.points[0]
        trace_idx = pt.get("curve_number", -1)
        yr_idx = trace_idx - YEAR_TRACE_OFFSET
        if 0 <= yr_idx < len(YEARS):
            selected_yr = YEARS[yr_idx]

    if selected_yr:
        _scenario_drill(M, selected_yr)
    else:
        st.info("☝️ Click a year dot on the map above to open the detailed analysis.")


def _scenario_drill(M, yr):
    """Detailed drill-down for a selected year's scenario."""
    case_num = M[yr]["Scenario"]
    label, icon, color, signs, interp = SCENARIOS[case_num]

    st.markdown(
        f'<div class="drill-panel">'
        f'<div style="font-size:1.2rem;font-weight:700;color:{color};">'
        f'{icon} FY {yr} — Case {case_num}: {label}</div>'
        f'<div style="font-size:.8rem;color:{TXT2};margin:4px 0;">{signs}</div>'
        f'<div style="font-size:.88rem;color:{TXT};margin-top:6px;">{interp}</div>'
        f'</div>', unsafe_allow_html=True)

    c1, c2, c3 = st.columns(3)
    with c1:
        wc_v = M[yr]["WC"]
        _, wc_c, _ = rate_sign(wc_v)
        card_html("Working Capital", fm(wc_v), wc_c,
                  f"= Equity + NCL − NCA")
    with c2:
        wcn_v = M[yr]["WCN"]
        _, wcn_c, _ = rate_sign(wcn_v)
        card_html("Working Capital Need", fm(wcn_v), wcn_c,
                  f"= (Inv + AR + OCA) − (AP + OCL)")
    with c3:
        nc_v = M[yr]["NC"]
        _, nc_c, _ = rate_sign(nc_v)
        card_html("Net Cash", fm(nc_v), nc_c,
                  f"= WC − WCN")

    # YoY movement vs previous year
    if yr != YEARS[0]:
        prev_yr = YEARS[YEARS.index(yr) - 1]
        st.markdown(f"**What changed vs {prev_yr}:**")
        deltas = {}
        for key in ["WC", "WCN", "NC", "Net Debt"]:
            deltas[key] = yoy_pct(M[yr][key], M[prev_yr][key])

        delta_cols = st.columns(4)
        for i, (key, chg) in enumerate(deltas.items()):
            with delta_cols[i]:
                st.markdown(
                    f'<div style="background:{BG};border:1px solid {BORDER};border-radius:8px;'
                    f'padding:8px 10px;text-align:center;">'
                    f'<div style="font-size:.72rem;color:{TXT2};">{key}</div>'
                    f'<div style="font-size:.95rem;">{arrow_html(chg)}</div>'
                    f'<div style="font-size:.68rem;color:{TXT2};">{fm(M[yr][key])} vs {fm(M[prev_yr][key])}</div>'
                    f'</div>', unsafe_allow_html=True)

        prev_case = M[prev_yr]["Scenario"]
        if prev_case != case_num:
            p_label,p_icon,p_color,_,_ = SCENARIOS[prev_case]
            st.markdown(
                f'<div style="margin-top:10px;padding:10px;background:{BG};'
                f'border:1px solid {BORDER};border-radius:8px;">'
                f'<span style="color:{TXT2};font-size:.82rem;">Scenario moved: </span>'
                f'<span style="color:{p_color};font-weight:700;">{p_icon} Case {prev_case}: {p_label}</span>'
                f' <span style="color:{TXT2};">→</span> '
                f'<span style="color:{color};font-weight:700;">{icon} Case {case_num}: {label}</span>'
                f'</div>', unsafe_allow_html=True)


# =====================================================================
# KEY TAKEAWAYS — Dynamic 4-year narrative panels
# =====================================================================
def key_takeaways_section(M):
    section("📌 Key Takeaways — 4-Year Financial Story")
    c1, c2, c3 = st.columns(3)

    # ─ Liquidity ─
    with c1:
        cr  = {yr: M[yr]["Current Ratio"]  for yr in YEARS}
        qr  = {yr: M[yr]["Quick Ratio"]    for yr in YEARS}
        csr = {yr: M[yr]["Cash Ratio"]     for yr in YEARS}
        cr_t = _trend(cr)
        cr0_ico, cr0_lbl, _ = _yr_label_cr(cr[2021])
        cr4_ico, cr4_lbl, _ = _yr_label_cr(cr[2024])
        qr4_ico, qr4_lbl,  _ = _yr_label_qr(qr[2024])
        csr4_ico, csr4_lbl, _ = _yr_label_cashr(csr[2024])
        cr_chg = yoy_pct(cr[2024], cr[2021])
        cr_chg_str = f" ({abs(cr_chg):.1f}% {'↑' if cr_chg >= 0 else '↓'})" if cr_chg is not None else ""
        good_cr = GREEN if "improving" in cr_t else RED if "declining" in cr_t else ORANGE
        st.markdown(
            f'<div class="ipanel">'
            f'<div style="font-size:1rem;font-weight:700;color:{BLUE};margin-bottom:10px;">💧 Liquidity</div>'
            f'<div style="font-size:.84rem;line-height:1.7;">'
            f'The <b>Current Ratio</b> has been '
            f'<span style="color:{good_cr};font-weight:600;">{cr_t}</span>'
            f' over 2021–2024{cr_chg_str}, moving from {cr0_ico} <b>{cr0_lbl}</b> ({fr(cr[2021])}) '
            f'to {cr4_ico} <b>{cr4_lbl}</b> ({fr(cr[2024])}) in 2024.<br><br>'
            f'In 2024, the <b>Quick Ratio</b> stands at {fr(qr[2024])} — {qr4_ico} <b>{qr4_lbl}</b> — '
            f'and the <b>Cash Ratio</b> at {fr(csr[2024])} — {csr4_ico} <b>{csr4_lbl}</b>.'
            f'</div></div>', unsafe_allow_html=True)

    # ─ Solvency & Leverage ─
    with c2:
        tdr = {yr: M[yr]["Total Debt Ratio"]  for yr in YEARS}
        de  = {yr: M[yr]["Debt-to-Equity"]    for yr in YEARS}
        em  = {yr: M[yr]["Equity Multiplier"] for yr in YEARS}
        tdr_t = _trend(tdr)
        tdr0_ico, tdr0_lbl, _ = _yr_label_tdr(tdr[2021])
        tdr4_ico, tdr4_lbl, _ = _yr_label_tdr(tdr[2024])
        de4_ico,  de4_lbl,  _ = _yr_label_de(de[2024])
        em4_ico,  em4_lbl,  _ = _yr_label_em(em[2024])
        tdr_chg = yoy_pct(tdr[2024], tdr[2021])
        tdr_chg_str = f" ({abs(tdr_chg):.1f}% {'↑' if tdr_chg >= 0 else '↓'})" if tdr_chg is not None else ""
        # lower TDR is better → reverse color logic
        good_tdr = GREEN if "declining" in tdr_t else RED if "improving" in tdr_t else ORANGE
        st.markdown(
            f'<div class="ipanel">'
            f'<div style="font-size:1rem;font-weight:700;color:{BLUE};margin-bottom:10px;">🏦 Solvency & Leverage</div>'
            f'<div style="font-size:.84rem;line-height:1.7;">'
            f'The <b>Total Debt Ratio</b> has been '
            f'<span style="color:{good_tdr};font-weight:600;">{tdr_t}</span>'
            f' over 2021–2024{tdr_chg_str}, from {tdr0_ico} <b>{tdr0_lbl}</b> ({fr(tdr[2021])}) '
            f'to {tdr4_ico} <b>{tdr4_lbl}</b> ({fr(tdr[2024])}) in 2024.<br><br>'
            f'In 2024, the <b>D/E ratio</b> of {fr(de[2024])} signals {de4_ico} <b>{de4_lbl}</b>, '
            f'and the <b>Equity Multiplier</b> of {fr(em[2024])} indicates {em4_ico} <b>{em4_lbl}</b>.'
            f'</div></div>', unsafe_allow_html=True)

    # ─ Financial Structure ─
    with c3:
        wc  = {yr: M[yr]["WC"]       for yr in YEARS}
        wcn = {yr: M[yr]["WCN"]      for yr in YEARS}
        nc  = {yr: M[yr]["NC"]       for yr in YEARS}
        nd  = {yr: M[yr]["Net Debt"] for yr in YEARS}
        wc4_ico,  wc4_lbl,  _ = _yr_label_wc(wc[2024])
        wcn4_ico, wcn4_lbl, _ = _yr_label_wcn(wcn[2024])
        nc4_ico,  nc4_lbl,  _ = _yr_label_nc(nc[2024])
        nd4_ico,  nd4_lbl,  _ = _yr_label_nd(nd[2024])
        c21, c24 = M[2021]["Scenario"], M[2024]["Scenario"]
        lbl21, lbl24 = SCENARIOS[c21][0], SCENARIOS[c24][0]
        scenario_path = " → ".join(f"Case {M[yr]['Scenario']}" for yr in YEARS)
        st.markdown(
            f'<div class="ipanel">'
            f'<div style="font-size:1rem;font-weight:700;color:{BLUE};margin-bottom:10px;">🏗️ Financial Structure</div>'
            f'<div style="font-size:.84rem;line-height:1.7;">'
            f'Adidas followed the scenario path <b>{scenario_path}</b>, from '
            f'<b>Case {c21} — {lbl21}</b> in 2021 to <b>Case {c24} — {lbl24}</b> in 2024.<br><br>'
            f'In 2024: WC {wc4_ico} <b>{wc4_lbl}</b> ({fm(wc[2024])}), '
            f'WCN {wcn4_ico} <b>{wcn4_lbl}</b> ({fm(wcn[2024])}), '
            f'NC {nc4_ico} <b>{nc4_lbl}</b> ({fm(nc[2024])}), '
            f'and {nd4_ico} <b>{nd4_lbl}</b> of {fm(abs(nd[2024]))}.'
            f'</div></div>', unsafe_allow_html=True)


# =====================================================================
# PAGE 1 — Executive Summary
# =====================================================================
def page_summary(df, M):
    st.title("👟 Adidas — Executive Summary")

    section("Health Snapshot by Year")
    cols = st.columns(4)
    for i, yr in enumerate(YEARS):
        with cols[i]:
            scenario_card_html(yr, M[yr]["Scenario"])

    # ── ★ Row 2: interactive expander cards for 2024 KPIs ──
    section("Key Metrics — FY 2024  (click any card to expand all years + trend)")
    cols = st.columns(4)

    with cols[0]:
        metric_card_expander(
            "Working Capital", {yr: M[yr]["WC"] for yr in YEARS},
            is_eur=True,
            formula="WC = Total Equity + Total NCL − Total NCA",
            impacts=["Net Cash (NC)", "Scenario classification"],
            key_prefix="sum_wc",
            interpret_fn=lambda v: _interp_block(v, _yr_label_wc, "Working Capital", higher_is_better=True),
        )
    with cols[1]:
        metric_card_expander(
            "Net Cash", {yr: M[yr]["NC"] for yr in YEARS},
            is_eur=True,
            formula="NC = WC − WCN  [verify: Cash − Borr.Current − Overdraft]",
            impacts=["Scenario classification (sign determines case)"],
            key_prefix="sum_nc",
            interpret_fn=lambda v: _interp_block(v, _yr_label_nc, "Net Cash", higher_is_better=True),
        )
    with cols[2]:
        metric_card_expander(
            "Current Ratio", {yr: M[yr]["Current Ratio"] for yr in YEARS},
            is_eur=False,
            formula="Current Ratio = Total Current Assets / Total Current Liabilities",
            threshold_html=(
                f'<span style="color:{RED};">● &lt;1.0 — Concern</span> &nbsp;'
                f'<span style="color:{GREEN};">● 1.5–2.0 — Healthy</span> &nbsp;'
                f'<span style="color:{ORANGE};">● &gt;3.0 — Inefficient</span>'
            ),
            impacts=["Liquidity assessment"],
            key_prefix="sum_cr",
            rate_fn=rate_cr,
            interpret_fn=lambda v: _interp_block(v, _yr_label_cr, "the Current Ratio", higher_is_better=True),
        )
    with cols[3]:
        metric_card_expander(
            "Debt-to-Equity", {yr: M[yr]["Debt-to-Equity"] for yr in YEARS},
            is_eur=False,
            formula="D/E = (Total NCL + Borrowings Current) / Total Equity",
            threshold_html=(
                f'<span style="color:{GREEN};">● 0.5–1.0 — Healthy</span> &nbsp;'
                f'<span style="color:{ORANGE};">● 1.0–1.5 — Acceptable</span> &nbsp;'
                f'<span style="color:{RED};">● &gt;1.5 — High leverage</span>'
            ),
            impacts=["Solvency assessment", "Invested Capital"],
            key_prefix="sum_de",
            rate_fn=rate_de,
            interpret_fn=lambda v: _interp_block(v, _yr_label_de, "Debt-to-Equity", higher_is_better=False),
        )

    key_takeaways_section(M)

    section("WC / WCN / NC Evolution (€M)")
    fig = go.Figure()
    for name, key, color in [("Working Capital","WC",BLUE),("WCN","WCN",ORANGE),("Net Cash","NC",GREEN)]:
        fig.add_trace(go.Scatter(
            x=YEARS_STR, y=[M[y][key]/1000 for y in YEARS],
            name=name, mode="lines+markers",
            line=dict(color=color,width=3), marker=dict(size=9)
        ))
    fig.add_hline(y=0,line_dash="dash",line_color=TXT2,opacity=0.4)
    fig.update_layout(yaxis_title="€M",xaxis_title="Year")
    st.plotly_chart(dark_fig(fig), use_container_width=True)

    section("All Computed Metrics")
    agg_keys=["WC","WCN","NC","Net Debt","Capital Employed","Invested Capital"]
    liq_keys=["Current Ratio","Quick Ratio","Cash Ratio"]
    sol_keys=["Equity Multiplier","Total Debt Ratio","Debt-to-Equity"]

    def _color_for(k,v):
        if k=="Current Ratio":    _,c,_=rate_cr(v);    return c
        if k=="Quick Ratio":      _,c,_=rate_qr(v);    return c
        if k=="Cash Ratio":       _,c,_=rate_cashr(v); return c
        if k=="Total Debt Ratio": _,c,_=rate_tdr(v);   return c
        if k=="Debt-to-Equity":   _,c,_=rate_de(v);    return c
        if k=="Equity Multiplier": return BLUE
        _,c,_=rate_sign(v); return c

    def _rows(keys, fmt):
        h=""
        for k in keys:
            h+=f'<tr><td style="padding:6px 12px;border-bottom:1px solid {BORDER};">{k}</td>'
            for yr in YEARS:
                v=M[yr][k]; c=_color_for(k,v)
                h+=(f'<td style="padding:6px 12px;border-bottom:1px solid {BORDER};'
                    f'text-align:right;color:{c};font-weight:600;">{fmt(v)}</td>')
            h+="</tr>"
        return h

    hdr=(f'<tr><th style="padding:8px 12px;text-align:left;border-bottom:2px solid {BLUE};">Metric</th>'
         +"".join(f'<th style="padding:8px 12px;text-align:right;border-bottom:2px solid {BLUE};">{y}</th>' for y in YEARS)
         +"</tr>")
    sr=lambda t:(f'<tr><td colspan="{len(YEARS)+1}" style="padding:8px 12px;color:{BLUE};'
                 f'font-weight:700;background:{BG};">{t}</td></tr>')
    st.markdown(
        f'<table style="width:100%;border-collapse:collapse;background:{CARD_BG};'
        f'border-radius:10px;overflow:hidden;font-size:.88rem;">'
        f'{hdr}{sr("Structural Aggregates (€M)")}{_rows(agg_keys,fm)}'
        f'{sr("Liquidity Ratios")}{_rows(liq_keys,fr)}'
        f'{sr("Solvency &amp; Leverage")}{_rows(sol_keys,fr)}</table>',
        unsafe_allow_html=True)


# =====================================================================
# PAGE 2 — Balance Sheet Explorer (★ drillable rows)
# =====================================================================
def page_balance_sheet(df, M):
    st.title("📊 Balance Sheet Explorer")

    section("Full Balance Sheet — Click a Row to Drill Down")
    st.caption("Select any row in the table to open the detailed drill panel below.")

    # Build numeric DataFrame for interactive selection
    rows_data = {}
    for item in ALL_ITEMS:
        if item in df.index:
            rows_data[item] = {yr: _g(df, item, yr)/1000 for yr in YEARS}
    numeric_df = pd.DataFrame(rows_data).T
    numeric_df.columns = [str(yr) for yr in YEARS]
    numeric_df.index.name = "Line Item"

    # Render with clickable rows
    col_cfg = {str(yr): st.column_config.NumberColumn(
        f"FY {yr}", format="€%.1fM") for yr in YEARS}

    event = st.dataframe(
        numeric_df,
        use_container_width=True,
        column_config=col_cfg,
        on_select="rerun",
        selection_mode="single-row",
        key="bs_table_interactive",
        height=520,
    )

    # Drill panel
    selected_item = None
    if event and hasattr(event, "selection") and event.selection.rows:
        row_idx = event.selection.rows[0]
        selected_item = numeric_df.index[row_idx]

    if selected_item:
        bs_drill_panel(df, M, selected_item)
    else:
        st.info("☝️ Click any row in the table above to open the detailed drill panel.")

    # ── Composition charts ──
    section("Asset Composition (€M)")
    asset_detail = [i for i in ASSET_ITEMS if not i.startswith("Total")]
    fig = go.Figure()
    for idx, item in enumerate(asset_detail):
        fig.add_trace(go.Bar(x=YEARS_STR, y=[_g(df,item,yr)/1000 for yr in YEARS],
                              name=sn(item), marker_color=PALETTE[idx%len(PALETTE)]))
    fig.update_layout(barmode="stack",yaxis_title="€M",title="Asset Breakdown by Year")
    st.plotly_chart(dark_fig(fig), use_container_width=True)

    section("Equity & Liabilities Composition (€M)")
    el_detail = [i for i in EQLIAB_ITEMS
                 if not i.startswith("Total") and any(_g(df,i,yr)!=0 for yr in YEARS)]
    fig = go.Figure()
    for idx, item in enumerate(el_detail):
        fig.add_trace(go.Bar(x=YEARS_STR, y=[_g(df,item,yr)/1000 for yr in YEARS],
                              name=sn(item), marker_color=PALETTE[idx%len(PALETTE)]))
    fig.update_layout(barmode="stack",yaxis_title="€M",title="E&L Breakdown by Year")
    st.plotly_chart(dark_fig(fig), use_container_width=True)

    section("Structural Breakdown (%)")
    c1, c2 = st.columns(2)
    with c1:
        fig = go.Figure()
        fig.add_trace(go.Bar(x=YEARS_STR,
            y=[_g(df,"Total non-current assets",yr)/_g(df,"Total assets",yr)*100 for yr in YEARS],
            name="NCA %", marker_color=BLUE))
        fig.add_trace(go.Bar(x=YEARS_STR,
            y=[_g(df,"Total current assets",yr)/_g(df,"Total assets",yr)*100 for yr in YEARS],
            name="CA %", marker_color=GREEN))
        fig.update_layout(barmode="stack",yaxis_title="%",title="Assets: NCA vs CA")
        st.plotly_chart(dark_fig(fig,360), use_container_width=True)
    with c2:
        fig = go.Figure()
        for lbl,key,col in [("Equity %","Total equity",GREEN),
                              ("NCL %","Total non-current liabilities",BLUE),
                              ("CL %","Total current liabilities",ORANGE)]:
            pcts=[_g(df,key,yr)/_g(df,"Total equity and liabilities",yr)*100 for yr in YEARS]
            fig.add_trace(go.Bar(x=YEARS_STR,y=pcts,name=lbl,marker_color=col))
        fig.update_layout(barmode="stack",yaxis_title="%",title="E&L: Equity vs NCL vs CL")
        st.plotly_chart(dark_fig(fig,360), use_container_width=True)

    section("Each Line Item as % of Total Assets / Total E&L")
    hdr2=(f'<tr><th style="padding:5px 12px;text-align:left;border-bottom:2px solid {BLUE};">Item</th>'
          +"".join(f'<th style="padding:5px 12px;text-align:right;border-bottom:2px solid {BLUE};">{y}</th>' for y in YEARS)
          +"</tr>")
    prows=""
    for item in ALL_ITEMS:
        if item not in df.index or item in ("Total assets","Total equity and liabilities"): continue
        dk="Total assets" if item in ASSET_ITEMS else "Total equity and liabilities"
        is_t=item.startswith("Total"); fw="font-weight:700;" if is_t else ""; bg=f"background:{BG};" if is_t else ""
        prows+=f'<tr style="{bg}"><td style="padding:4px 12px;border-bottom:1px solid {BORDER};{fw}">{sn(item)}</td>'
        for yr in YEARS:
            d=_g(df,dk,yr); p=_g(df,item,yr)/d*100 if d else 0
            prows+=f'<td style="padding:4px 12px;border-bottom:1px solid {BORDER};text-align:right;{fw}">{p:.1f}%</td>'
        prows+="</tr>"
    st.markdown(f'<table style="width:100%;border-collapse:collapse;background:{CARD_BG};'
                f'border-radius:10px;overflow:hidden;font-size:.82rem;">{hdr2}{prows}</table>',
                unsafe_allow_html=True)

    section("Year-over-Year Growth (%)")
    hdr3=(f'<tr><th style="padding:5px 12px;text-align:left;border-bottom:2px solid {BLUE};">Item</th>'
          +"".join(f'<th style="padding:5px 12px;text-align:right;border-bottom:2px solid {BLUE};">{y}</th>' for y in YEARS[1:])
          +"</tr>")
    grows=""
    for item in ALL_ITEMS:
        if item not in df.index: continue
        is_t=item.startswith("Total"); fw="font-weight:700;" if is_t else ""; bg=f"background:{BG};" if is_t else ""
        grows+=f'<tr style="{bg}"><td style="padding:4px 12px;border-bottom:1px solid {BORDER};{fw}">{sn(item)}</td>'
        for i in range(1,len(YEARS)):
            chg=yoy_pct(_g(df,item,YEARS[i]),_g(df,item,YEARS[i-1]))
            grows+=f'<td style="padding:4px 12px;border-bottom:1px solid {BORDER};text-align:right;{fw}">{arrow_html(chg)}</td>'
        grows+="</tr>"
    st.markdown(f'<table style="width:100%;border-collapse:collapse;background:{CARD_BG};'
                f'border-radius:10px;overflow:hidden;font-size:.82rem;">{hdr3}{grows}</table>',
                unsafe_allow_html=True)


# =====================================================================
# PAGE 3 — Structural Aggregates (★ scenario scatter + drill)
# =====================================================================
def page_aggregates(df, M):
    st.title("🏗️ Structural Aggregates")
    AGG = ["WC","WCN","NC","Net Debt","Capital Employed","Invested Capital"]

    # ★ Scenario scatter at the top — the signature interactive feature
    scenario_scatter_section(M)

    section("Formulas")
    st.markdown(
        f'<div class="ipanel" style="font-size:.88rem;">'
        f'<div class="formula">WC  = Total Equity + Total NCL − Total NCA</div>'
        f'<div class="formula">WCN = (Inventories + Accounts Receivable + Other CA) − (Accounts Payable + Other CL)</div>'
        f'<div class="formula">NC  = WC − WCN &nbsp;&nbsp;[verify: Cash − Borrowings Current − Bank Overdraft]</div>'
        f'<div class="formula">Net Debt = Total NCL + Borrowings Current − Cash</div>'
        f'<div class="formula">Capital Employed = Total NCA + WCN</div>'
        f'<div class="formula">Invested Capital = Total Equity + Net Debt</div>'
        f'</div>', unsafe_allow_html=True)

    # ★ Interactive aggregate cards with expanders
    section("Aggregate Values — Click Any Card to Expand All Years & Trend")

    FORMULAS = {
        "WC":               "Total Equity + Total NCL − Total NCA",
        "WCN":              "(Inventories + AR + Other CA) − (AP + Other CL)",
        "NC":               "WC − WCN  [verify: Cash − Borr.Current − Overdraft]",
        "Net Debt":         "Total NCL + Borrowings Current − Cash",
        "Capital Employed": "Total NCA + WCN",
        "Invested Capital": "Total Equity + Net Debt",
    }
    AGG_IMPACTS = {
        "WC":               ["Scenario classification (sign)","Net Cash"],
        "WCN":              ["Net Cash","Scenario classification","Capital Employed"],
        "NC":               ["Scenario classification (sign determines case)"],
        "Net Debt":         ["Invested Capital","Debt-to-Equity","Total Debt Ratio"],
        "Capital Employed": ["Invested Capital"],
        "Invested Capital": ["Benchmark for capital efficiency"],
    }

    AGG_INTERP = {
        "WC":               lambda v: _interp_block(v, _yr_label_wc,  "Working Capital",      higher_is_better=True),
        "WCN":              lambda v: _interp_block(v, _yr_label_wcn, "Working Capital Need",  higher_is_better=False),
        "NC":               lambda v: _interp_block(v, _yr_label_nc,  "Net Cash",              higher_is_better=True),
        "Net Debt":         lambda v: _interp_block(v, _yr_label_nd,  "Net Debt",              higher_is_better=False),
        "Capital Employed": lambda v: _interp_block(v, _yr_label_ce,  "Capital Employed",      higher_is_better=True),
        "Invested Capital": lambda v: _interp_block(v, _yr_label_ic,  "Invested Capital",      higher_is_better=True),
    }
    c1, c2, c3 = st.columns(3)
    for i, key in enumerate(AGG):
        col = [c1, c2, c3][i % 3]
        with col:
            metric_card_expander(
                key, {yr: M[yr][key] for yr in YEARS},
                is_eur=True, formula=FORMULAS.get(key,""),
                impacts=AGG_IMPACTS.get(key,[]),
                key_prefix=f"agg_{key}",
                interpret_fn=AGG_INTERP.get(key),
            )

    # NC verification
    st.caption("NC verification per year: " +
               " | ".join(f"{yr}: {fm(M[yr]['NC_verify'])}" for yr in YEARS))

    section("Scenario Classification")
    cols = st.columns(4)
    for i, yr in enumerate(YEARS):
        with cols[i]:
            scenario_card_html(yr, M[yr]["Scenario"])

    section("Aggregates Comparison (€M)")
    fig = go.Figure()
    for idx, key in enumerate(AGG):
        fig.add_trace(go.Bar(x=YEARS_STR,y=[M[yr][key]/1000 for yr in YEARS],
                              name=key,marker_color=PALETTE[idx%len(PALETTE)]))
    fig.update_layout(barmode="group",yaxis_title="€M")
    fig.add_hline(y=0,line_dash="dash",line_color=TXT2,opacity=0.4)
    st.plotly_chart(dark_fig(fig,480), use_container_width=True)

    section("Year-over-Year Growth")
    hdr=(f'<tr><th style="padding:6px 12px;text-align:left;border-bottom:2px solid {BLUE};">Aggregate</th>'
         +"".join(f'<th style="padding:6px 12px;text-align:right;border-bottom:2px solid {BLUE};">{y}</th>' for y in YEARS[1:])
         +"</tr>")
    rows_h=""
    for key in AGG:
        rows_h+=f'<tr><td style="padding:5px 12px;border-bottom:1px solid {BORDER};font-weight:600;">{key}</td>'
        for i in range(1,len(YEARS)):
            chg=yoy_pct(M[YEARS[i]][key],M[YEARS[i-1]][key])
            rows_h+=f'<td style="padding:5px 12px;border-bottom:1px solid {BORDER};text-align:right;">{arrow_html(chg)}</td>'
        rows_h+="</tr>"
    st.markdown(f'<table style="width:100%;border-collapse:collapse;background:{CARD_BG};'
                f'border-radius:10px;overflow:hidden;font-size:.86rem;">{hdr}{rows_h}</table>',
                unsafe_allow_html=True)


# =====================================================================
# PAGE 4 — Liquidity Analysis (★ expander interpretation panels)
# =====================================================================
def _ratio_chart(M, key, title, thresholds, rate_fn):
    fig = go.Figure()
    colors = [rate_fn(M[yr][key])[1] for yr in YEARS]
    vals = [M[yr][key] for yr in YEARS]
    fig.add_trace(go.Bar(x=YEARS_STR,y=vals,marker_color=colors,name=title,
                          text=[fr(v) for v in vals],textposition="outside",
                          textfont=dict(color=TXT)))
    for level,lbl,col,dash in thresholds:
        fig.add_hline(y=level,line_dash=dash,line_color=col,opacity=0.6,
                       annotation_text=lbl,annotation_position="top left",
                       annotation_font_color=col,annotation_font_size=11)
    fig.update_layout(yaxis_title="Ratio",title=title,showlegend=False)
    return dark_fig(fig)


def page_liquidity(df, M):
    st.title("💧 Liquidity Analysis")

    for key, title, formula, thresh_html, thresh_lines, rate_fn in [
        ("Current Ratio","Current Ratio",
         "Total Current Assets / Total Current Liabilities",
         (f'<span style="color:{RED};">● &lt;1.0 Concern</span> &nbsp;'
          f'<span style="color:{GREEN};">● 1.5–2.0 Healthy</span> &nbsp;'
          f'<span style="color:{ORANGE};">● &gt;3.0 Inefficient</span>'),
         [(1,"1.0 Critical",RED,"dash"),(1.5,"1.5 Healthy min",GREEN,"dot"),(2,"2.0 Healthy max",GREEN,"dot")],
         rate_cr),
        ("Quick Ratio","Quick Ratio",
         "(Total Current Assets − Inventories) / Total Current Liabilities",
         (f'<span style="color:{RED};">● &lt;1.0 Risk</span> &nbsp;'
          f'<span style="color:{GREEN};">● 1.0–1.5 Strong</span> &nbsp;'
          f'<span style="color:{ORANGE};">● &gt;2.0 Conservative</span>'),
         [(1,"1.0 Risk",RED,"dash"),(1.5,"1.5 Strong",GREEN,"dot")],
         rate_qr),
        ("Cash Ratio","Cash Ratio",
         "Cash and Cash Equivalents / Total Current Liabilities",
         (f'<span style="color:{ORANGE};">● &lt;0.5 Relies on inv./recv.</span> &nbsp;'
          f'<span style="color:{GREEN};">● 0.5–1.0 Ideal</span> &nbsp;'
          f'<span style="color:{ORANGE};">● &gt;1.0 Conservative</span>'),
         [(0.5,"0.5 Min ideal",ORANGE,"dash"),(1,"1.0 Max ideal",GREEN,"dot")],
         rate_cashr),
    ]:
        section(title)
        c1, c2 = st.columns([1, 2])
        with c1:
            # ★ Interactive expander card (all years + trend inside)
            LIQ_INTERP = {
                "Current Ratio": lambda v: _interp_block(v, _yr_label_cr,    "the Current Ratio", higher_is_better=True),
                "Quick Ratio":   lambda v: _interp_block(v, _yr_label_qr,    "the Quick Ratio",   higher_is_better=True),
                "Cash Ratio":    lambda v: _interp_block(v, _yr_label_cashr, "the Cash Ratio",    higher_is_better=True),
            }
            metric_card_expander(
                title, {yr: M[yr][key] for yr in YEARS},
                is_eur=False, formula=formula,
                threshold_html=thresh_html, key_prefix=f"liq_{key}",
                rate_fn=rate_fn,
                interpret_fn=LIQ_INTERP.get(key),
            )
            # Verdict for 2024
            e, co, sub = rate_fn(M[2024][key])
            st.markdown(
                f'<div class="ipanel">'
                f'<div class="verdict" style="background:{co}18;border-left:3px solid {co};color:{co};">'
                f'{e} 2024: {sub}</div></div>', unsafe_allow_html=True)
        with c2:
            st.plotly_chart(_ratio_chart(M, key, title, thresh_lines, rate_fn),
                            use_container_width=True)

    section("Comparative — All Liquidity Ratios")
    fig = go.Figure()
    for name, key, col in [("Current Ratio","Current Ratio",BLUE),
                             ("Quick Ratio","Quick Ratio",GREEN),
                             ("Cash Ratio","Cash Ratio",ORANGE)]:
        fig.add_trace(go.Scatter(x=YEARS_STR,y=[M[yr][key] for yr in YEARS],
                                  name=name,mode="lines+markers",
                                  line=dict(color=col,width=3),marker=dict(size=9)))
    fig.add_hline(y=1,line_dash="dash",line_color=RED,opacity=0.4,
                   annotation_text="1.0",annotation_font_color=RED)
    fig.update_layout(yaxis_title="Ratio",xaxis_title="Year")
    st.plotly_chart(dark_fig(fig), use_container_width=True)


# =====================================================================
# PAGE 5 — Solvency & Leverage (★ expander interpretation panels)
# =====================================================================
def page_solvency(df, M):
    st.title("🏦 Solvency & Leverage")

    # Equity Multiplier
    section("Equity Multiplier")
    c1, c2 = st.columns([1, 2])
    with c1:
        metric_card_expander(
            "Equity Multiplier", {yr: M[yr]["Equity Multiplier"] for yr in YEARS},
            is_eur=False,
            formula="Total Assets / Total Equity",
            threshold_html=(
                f'<span style="color:{TXT};">EM = 1 → 100% equity financed.<br>'
                f'Higher EM = more debt-financed = more leveraged.</span>'
            ),
            impacts=["Financial leverage aggressiveness"],
            key_prefix="sol_em",
            interpret_fn=lambda v: _interp_block(v, _yr_label_em, "the Equity Multiplier", higher_is_better=False),
        )
    with c2:
        st.plotly_chart(_ratio_chart(M,"Equity Multiplier","Equity Multiplier",[],
                                      lambda v:("",BLUE,"")),
                        use_container_width=True)

    # Total Debt Ratio
    section("Total Debt Ratio")
    c1, c2 = st.columns([1, 2])
    with c1:
        e, co, sub = rate_tdr(M[2024]["Total Debt Ratio"])
        metric_card_expander(
            "Total Debt Ratio", {yr: M[yr]["Total Debt Ratio"] for yr in YEARS},
            is_eur=False,
            formula="(Total NCL + Borrowings Current) / Total Assets",
            threshold_html=(
                f'<span style="color:{GREEN};">● 0.2–0.4 Low debt</span><br>'
                f'<span style="color:{ORANGE};">● 0.4–0.6 Acceptable</span><br>'
                f'<span style="color:{RED};">● &gt;0.6 High &nbsp; · &nbsp; &gt;0.8 Very high</span>'
            ),
            impacts=["Solvency assessment","Debt capacity"],
            key_prefix="sol_tdr", rate_fn=rate_tdr,
            interpret_fn=lambda v: _interp_block(v, _yr_label_tdr, "the Total Debt Ratio", higher_is_better=False),
        )
        st.markdown(
            f'<div class="ipanel">'
            f'<div class="verdict" style="background:{co}18;border-left:3px solid {co};color:{co};">'
            f'{e} 2024: {sub}</div></div>', unsafe_allow_html=True)
    with c2:
        thresholds=[(0.4,"0.4 Low max",GREEN,"dot"),(0.6,"0.6 Acceptable max",ORANGE,"dash"),(0.8,"0.8 High",RED,"dash")]
        st.plotly_chart(_ratio_chart(M,"Total Debt Ratio","Total Debt Ratio",thresholds,rate_tdr),
                        use_container_width=True)

    # D/E
    section("Debt-to-Equity")
    c1, c2 = st.columns([1, 2])
    with c1:
        e, co, sub = rate_de(M[2024]["Debt-to-Equity"])
        metric_card_expander(
            "Debt-to-Equity", {yr: M[yr]["Debt-to-Equity"] for yr in YEARS},
            is_eur=False,
            formula="(Total NCL + Borrowings Current) / Total Equity",
            threshold_html=(
                f'<span style="color:{GREEN};">● 0.5–1.0 Healthy</span><br>'
                f'<span style="color:{ORANGE};">● 1.0–1.5 Acceptable</span><br>'
                f'<span style="color:{RED};">● &gt;1.5 High leverage</span>'
            ),
            impacts=["Leverage","Solvency risk","Invested Capital"],
            key_prefix="sol_de", rate_fn=rate_de,
            interpret_fn=lambda v: _interp_block(v, _yr_label_de, "Debt-to-Equity", higher_is_better=False),
        )
        st.markdown(
            f'<div class="ipanel">'
            f'<div class="verdict" style="background:{co}18;border-left:3px solid {co};color:{co};">'
            f'{e} 2024: {sub}</div></div>', unsafe_allow_html=True)
    with c2:
        thresholds=[(1,"1.0 Healthy max",GREEN,"dot"),(1.5,"1.5 Acceptable max",ORANGE,"dash")]
        st.plotly_chart(_ratio_chart(M,"Debt-to-Equity","Debt-to-Equity",thresholds,rate_de),
                        use_container_width=True)

    section("Comparative — Solvency Ratios")
    fig = go.Figure()
    for name, key, col in [("Equity Multiplier","Equity Multiplier",BLUE),
                             ("Total Debt Ratio","Total Debt Ratio",ORANGE),
                             ("Debt-to-Equity","Debt-to-Equity",GREEN)]:
        fig.add_trace(go.Scatter(x=YEARS_STR,y=[M[yr][key] for yr in YEARS],
                                  name=name,mode="lines+markers",
                                  line=dict(color=col,width=3),marker=dict(size=9)))
    fig.update_layout(yaxis_title="Ratio",xaxis_title="Year")
    st.plotly_chart(dark_fig(fig), use_container_width=True)


# =====================================================================
# PAGE 6 — Interactive Comparison Tool
# =====================================================================
def page_comparison(df, M):
    st.title("🔀 Interactive Comparison Tool")
    eur_keys=["WC","WCN","NC","Net Debt","Capital Employed","Invested Capital"]
    ratio_keys=["Current Ratio","Quick Ratio","Cash Ratio","Equity Multiplier","Total Debt Ratio","Debt-to-Equity"]
    raw_items=[i for i in ALL_ITEMS if i in df.index]

    c1, c2 = st.columns(2)
    with c1: sel_years=st.multiselect("Select Years",YEARS,default=YEARS)
    with c2: metric_group=st.selectbox("Metric Group",
                ["Structural Aggregates (€M)","Liquidity & Solvency Ratios","Raw Balance Sheet Items (€M)"])
    if not sel_years: st.warning("Select at least one year."); return

    if metric_group=="Structural Aggregates (€M)":   avail,fmt,unit=eur_keys,fm,"€M"
    elif metric_group=="Liquidity & Solvency Ratios": avail,fmt,unit=ratio_keys,fr,"Ratio"
    else:                                              avail,fmt,unit=raw_items,fm,"€M"

    sel_metrics=st.multiselect("Select Metrics",avail,default=avail[:4])
    if not sel_metrics: st.warning("Select at least one metric."); return

    section("Comparison Chart")
    fig=go.Figure()
    for idx,key in enumerate(sel_metrics):
        if key in eur_keys or key in ratio_keys:
            vals=[M[yr][key]/1000 if key in eur_keys else M[yr][key] for yr in sel_years]
        else:
            vals=[_g(df,key,yr)/1000 for yr in sel_years]
        fig.add_trace(go.Bar(x=[str(y) for y in sel_years],y=vals,
                              name=sn(key),marker_color=PALETTE[idx%len(PALETTE)]))
    fig.update_layout(barmode="group",yaxis_title=unit,xaxis_title="Year")
    st.plotly_chart(dark_fig(fig,450), use_container_width=True)

    section("Values & Year-over-Year Change")
    hdr=f'<tr><th style="padding:6px 12px;text-align:left;border-bottom:2px solid {BLUE};">Metric</th>'
    for y in sel_years:
        hdr+=f'<th style="padding:6px 12px;text-align:right;border-bottom:2px solid {BLUE};">{y}</th>'
    for i in range(1,len(sel_years)):
        hdr+=f'<th style="padding:6px 12px;text-align:right;border-bottom:2px solid {BLUE};">Δ {sel_years[i-1]}→{sel_years[i]}</th>'
    hdr+="</tr>"
    trows=""
    for key in sel_metrics:
        trows+=f'<tr><td style="padding:5px 12px;border-bottom:1px solid {BORDER};font-weight:600;">{sn(key)}</td>'
        vals=[]
        for yr in sel_years:
            v=M[yr][key] if (key in eur_keys or key in ratio_keys) else _g(df,key,yr)
            vals.append(v)
            trows+=f'<td style="padding:5px 12px;border-bottom:1px solid {BORDER};text-align:right;">{fmt(v)}</td>'
        for i in range(1,len(sel_years)):
            chg=yoy_pct(vals[i],vals[i-1])
            trows+=f'<td style="padding:5px 12px;border-bottom:1px solid {BORDER};text-align:right;">{arrow_html(chg)}</td>'
        trows+="</tr>"
    st.markdown(f'<table style="width:100%;border-collapse:collapse;background:{CARD_BG};'
                f'border-radius:10px;overflow:hidden;font-size:.86rem;">{hdr}{trows}</table>',
                unsafe_allow_html=True)


# =====================================================================
# PAGE 7 — Trend & Forecast
# =====================================================================
def page_trends(df, M):
    st.title("📈 Trend & Forecast")
    st.caption("Linear regression on 4 data points (2021-2024). Projections are indicative only.")

    forecast_keys=[
        ("WC","Working Capital",fm,True),("NC","Net Cash",fm,True),
        ("Current Ratio","Current Ratio",fr,False),
        ("Net Debt","Net Debt",fm,True),("Debt-to-Equity","Debt-to-Equity",fr,False),
    ]
    x=np.array([0,1,2,3],dtype=float)
    x_proj=np.array([0,1,2,3,4],dtype=float)
    years_ext_str=YEARS_STR+["2025"]

    for key,title,fmt,is_eur in forecast_keys:
        section(f"{title} — Trend & 2025 Projection")
        y=np.array([M[yr][key] for yr in YEARS],dtype=float)
        y_d=y/1000 if is_eur else y.copy()
        coeffs=np.polyfit(x,y_d,1)
        y_fit=np.polyval(coeffs,x); y_proj=np.polyval(coeffs,4)
        y_fit_ext=np.polyval(coeffs,x_proj)
        ss_res=np.sum((y_d-y_fit)**2); ss_tot=np.sum((y_d-np.mean(y_d))**2)
        r2=1-ss_res/ss_tot if ss_tot>0 else 0

        c1, c2 = st.columns([1,2])
        with c1:
            card_html("2025 Projection",
                      fmt(y_proj*1000) if is_eur else fr(y_proj), BLUE,
                      f"R² = {r2:.3f}")
            r2_color=GREEN if r2>=0.5 else ORANGE
            r2_text=f"✅ R² = {r2:.3f} — Trend line fits the data reasonably well." if r2>=0.5 \
                    else f"⚠️ Low R² ({r2:.3f}) — Projection is unreliable."
            st.markdown(f'<div class="card" style="border-left:3px solid {r2_color};">'
                        f'<span style="color:{r2_color};font-weight:600;font-size:.88rem;">'
                        f'{r2_text}</span></div>', unsafe_allow_html=True)
        with c2:
            fig=go.Figure()
            fig.add_trace(go.Scatter(x=YEARS_STR,y=y_d.tolist(),name="Actual",mode="lines+markers",
                                      line=dict(color=BLUE,width=3),marker=dict(size=10)))
            fig.add_trace(go.Scatter(x=years_ext_str,y=y_fit_ext.tolist(),name="Trend",mode="lines",
                                      line=dict(color=GREEN,width=2,dash="dash")))
            fig.add_trace(go.Scatter(x=["2025"],y=[y_proj],name="2025 Projection",mode="markers",
                                      marker=dict(size=14,color=ORANGE,symbol="diamond")))
            fig.update_layout(yaxis_title="€M" if is_eur else "Ratio",
                               xaxis_title="Year",title=title)
            st.plotly_chart(dark_fig(fig), use_container_width=True)


# =====================================================================
# MAIN
# =====================================================================
PAGES = {
    "📋 Executive Summary":     page_summary,
    "📊 Balance Sheet Explorer": page_balance_sheet,
    "🏗️ Structural Aggregates":  page_aggregates,
    "💧 Liquidity Analysis":    page_liquidity,
    "🏦 Solvency & Leverage":   page_solvency,
    "🔀 Interactive Comparison": page_comparison,
    "📈 Trend & Forecast":      page_trends,
}

def main():
    inject_css()
    with st.sidebar:
        st.markdown(
            f'<div style="text-align:center;padding:12px 0;">'
            f'<div style="font-size:2rem;">👟</div>'
            f'<div style="font-size:1.2rem;font-weight:700;color:{TXT};">Adidas</div>'
            f'<div style="font-size:.75rem;color:{TXT2};">Financial Analysis Dashboard v2</div>'
            f'<div style="font-size:.65rem;color:{TXT2};margin-top:4px;">Balance Sheet 2021–2024</div>'
            f'</div>', unsafe_allow_html=True)
        st.markdown("---")
        st.markdown(
            f'<div style="font-size:.72rem;color:{TXT2};padding:0 8px 8px;">'
            f'<b style="color:{BLUE};">★ Interactive features</b><br>'
            f'• Click metric cards to expand<br>'
            f'• Click balance sheet rows to drill<br>'
            f'• Click year dots on the scenario map'
            f'</div>', unsafe_allow_html=True)
        page = st.radio("Navigation", list(PAGES.keys()), label_visibility="collapsed")

    if not FILE_PATH.exists():
        st.error(f"Data file not found:\n{FILE_PATH}")
        return

    df = load_data()
    M  = compute(df)
    PAGES[page](df, M)

if __name__ == "__main__":
    main()
